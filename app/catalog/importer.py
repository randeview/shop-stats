# app/catalog/importer.py
import os
import tempfile
from typing import Optional

from django.db import transaction
from openpyxl import load_workbook
from slugify import slugify

from .models import Category, Product


def _norm(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s or None


def import_categories_from_xlsx(file_obj, sheet_name: Optional[str] = None) -> int:
    """
    Simplest importer:
      - expects headers: PARENT_CATEGORY, CATEGORY_2LVL, CATEGORY_3LVL
      - for each row, does get_or_create for level1 -> level2 -> level3
      - returns number of categories created
    """
    created = 0

    # Save upload to a temp file (openpyxl needs a path-like)
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    try:
        with os.fdopen(fd, "wb") as tmp:
            if hasattr(file_obj, "chunks"):
                for chunk in file_obj.chunks():
                    tmp.write(chunk)
            else:
                tmp.write(file_obj.read())

        wb = load_workbook(tmp_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        # Parse header
        header = next(ws.iter_rows(values_only=True))
        header = [str(x).strip() if x else "" for x in header]
        try:
            p_idx = header.index("PARENT_CATEGORY")
            c2_idx = header.index("CATEGORY_2LVL")
            c3_idx = header.index("CATEGORY_3LVL")
            c4_idx = header.index("MERCHANT_NAME")
            c5_idx = header.index("PRODUCT")
            c6_idx = header.index("SKU")
            c7_idx = header.index("Кол-во товаров")
            c8_idx = header.index("Кол-во заказов")
        except ValueError:
            wb.close()
            raise ValueError(
                "Header must contain PARENT_CATEGORY, CATEGORY_2LVL, CATEGORY_3LVL"
            )

        # Import rows
        counter = 0
        with transaction.atomic():
            for row in ws.iter_rows(values_only=True):
                counter += 1
                lvl1 = _norm(row[p_idx])
                if not lvl1:
                    continue
                lvl2 = _norm(row[c2_idx])
                lvl3 = _norm(row[c3_idx])
                merchant_name = _norm(row[c4_idx])
                product_name = _norm(row[c5_idx])
                article_id = _norm(row[c6_idx])
                product_count = _norm(row[c7_idx])
                product_orders = _norm(row[c8_idx])

                if lvl1 == "PARENT_CATEGORY":
                    continue
                # Level 1
                c1, was_created = Category.objects.get_or_create(
                    parent=None,
                    slug=slugify(lvl1),
                    defaults={"name": lvl1},
                )
                if was_created:
                    created += 1

                # Level 2 (optional)
                if lvl2:
                    c2, was_created = Category.objects.get_or_create(
                        parent=c1,
                        slug=slugify(lvl2),
                        defaults={"name": lvl2},
                    )
                    if was_created:
                        created += 1
                else:
                    c2 = None

                # Level 3 (optional; only if level 2 exists)
                if lvl3 and c2:
                    c3, was_created = Category.objects.get_or_create(
                        parent=c2,
                        slug=slugify(lvl3),
                        defaults={"name": lvl3},
                    )
                    if was_created:
                        created += 1
                else:
                    c3 = Category.objects.get(name=lvl3)
                if not Product.objects.filter(
                    article_id=article_id, merchant_name=merchant_name
                ).exists():
                    product = Product.objects.create(
                        name=product_name,
                        category=c3,
                        merchant_name=merchant_name,
                        article_id=article_id,
                        product_count=product_count,
                        product_orders=product_orders,
                    )
                    product.save()
                if counter > 1000:
                    break

        wb.close()
        return created

    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
